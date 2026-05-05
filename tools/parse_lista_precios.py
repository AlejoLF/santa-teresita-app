#!/usr/bin/env python3
"""
SANTA TERESITA — Parser de "Lista de Precios.xlsx" → JSON de seed.

Lee el Excel del cliente y emite un JSON estructurado que el seed de Prisma usa
para poblar categorías, tipos de producto, productos y modificadores.

Uso:
    python tools/parse_lista_precios.py \
        --excel "Lista de Precios.xlsx" \
        --output packages/db/prisma/seed-data/lista-precios.json

Heurísticas (ver SPEC §2.9):
    - Hoja "Hoja 1" tiene 3 columnas: pastas (col 0-1), pizzas/tartas (col 2-3),
      porciones calientes (col 4-6).
    - Hoja "RESTO SIMPLE" tiene la jerarquía categoría → tipo → opciones de relleno.
      Una fila con col[0] no vacío es un nuevo TipoProducto. Filas con solo col[1]
      son opciones del tipo anterior.
    - "Hoja 1" contiene los precios actuales (16/04/2026) para todos los productos.
    - Hoja "Pedidos YA" tiene precios alternativos para canal Pedidos YA.

Salida JSON con shape:
    {
      "categorias": [{"nombre": "...", "orden": N}],
      "tipos_producto": [{"categoria": "...", "nombre": "...", "cocina_interviene": bool}],
      "productos": [{"tipo": "...", "nombre": "...", "precio_base": "...", ...}],
      "modificadores": [
        {"tipo_producto": "...", "grupo": "Relleno", "opciones": [...]}
      ],
      "precios_pedidos_ya": [{"producto_codigo": "...", "precio": "..."}]
    }
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Falta openpyxl. Instalá con: pip install openpyxl\n")
    sys.exit(1)


# ────────────────────────────────────────────────────────────────────────
#   Mapeo de categorías
# ────────────────────────────────────────────────────────────────────────

CATEGORIAS = [
    {"nombre": "Pastas frescas", "orden": 1, "icono": "🍝"},
    {"nombre": "Pizzas", "orden": 2, "icono": "🍕"},
    {"nombre": "Tartas", "orden": 3, "icono": "🥧"},
    {"nombre": "Salsas", "orden": 4, "icono": "🥫"},
    {"nombre": "Empanadas", "orden": 5, "icono": "🥟"},
    {"nombre": "Porciones calientes", "orden": 6, "icono": "🍽️"},
    {"nombre": "Otros", "orden": 7, "icono": "📦"},
]


# Mapeo de nombres del Excel → (categoría, tipo de producto, descripción)
# Lo que falta lo dejamos en "Otros" como fallback.
MAPEO_PRODUCTOS_HOJA1: dict[str, dict[str, Any]] = {
    # Pastas frescas
    "RAV VERDURA": {
        "categoria": "Pastas frescas",
        "tipo": "Ravioles",
        "forma_venta": "PLANCHA",
        "unidad_precio": "POR_PLANCHA",
        "cantidad_default": 1,
        "modificador_default": "Verdura",
    },
    "POLLO, CARNE": {
        "categoria": "Pastas frescas",
        "tipo": "Ravioles",
        "forma_venta": "PLANCHA",
        "unidad_precio": "POR_PLANCHA",
        "cantidad_default": 1,
        "modificador_default": "Pollo y Carne",
    },
    "RICOTA": {
        "categoria": "Pastas frescas",
        "tipo": "Ravioles",
        "forma_venta": "PLANCHA",
        "unidad_precio": "POR_PLANCHA",
        "cantidad_default": 1,
        "modificador_default": "Ricota",
    },
    "RIC JAMON": {
        "categoria": "Pastas frescas",
        "tipo": "Ravioles",
        "forma_venta": "PLANCHA",
        "unidad_precio": "POR_PLANCHA",
        "cantidad_default": 1,
        "modificador_default": "Ricota y Jamón",
    },
    "FID HUEVO": {
        "categoria": "Pastas frescas",
        "tipo": "Fideos al huevo",
        "forma_venta": "GRAMO",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 200,
    },
    "FID ESPEC": {
        "categoria": "Pastas frescas",
        "tipo": "Fideos especiales",
        "forma_venta": "GRAMO",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 200,
    },
    "ÑOQUIS": {
        "categoria": "Pastas frescas",
        "tipo": "Ñoquis de sémola",
        "forma_venta": "GRAMO",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 300,
    },
    "TORTELETTIS": {
        "categoria": "Pastas frescas",
        "tipo": "Tortelettis",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 300,
    },
    "SORRENTINOS": {
        "categoria": "Pastas frescas",
        "tipo": "Sorrentinos",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 300,
    },
    "SORRENTINOS DE SALMON": {
        "categoria": "Pastas frescas",
        "tipo": "Sorrentinos",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 300,
        "es_producto_separado": True,
        "nombre_separado": "Sorrentinos de Salmón",
    },
    "RAVIOLONES REMOLACHA": {
        "categoria": "Pastas frescas",
        "tipo": "Raviolones",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 300,
        "modificador_default": "Roquefort, Cebolla caramelizada y Nuez",
    },
    "LASAGNA": {
        "categoria": "Pastas frescas",
        "tipo": "Lasagna",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 1000,
    },
    "RONDELLI": {
        "categoria": "Pastas frescas",
        "tipo": "Rondelli",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 1000,
    },
    "CAN, VERDURA": {
        "categoria": "Pastas frescas",
        "tipo": "Canelones",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 4,
        "modificador_default": "Verdura",
    },
    "CAN, JAM Y QU": {
        "categoria": "Pastas frescas",
        "tipo": "Canelones",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 4,
        "modificador_default": "Jamón y Queso",
    },
    "CARNE Y VERDU": {
        "categoria": "Pastas frescas",
        "tipo": "Canelones",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 4,
        "modificador_default": "Carne y Verdura",
    },
    "FORATTI/FUCCILE/MOSTACHOLES": {
        "categoria": "Pastas frescas",
        "tipo": "Pasta seca",
        "forma_venta": "GRAMO",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 500,
    },
    "FORATTI / FUCCILE": {  # alias
        "categoria": "Pastas frescas",
        "tipo": "Pasta seca",
        "forma_venta": "GRAMO",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 500,
    },
    "PREPIZA": {
        "categoria": "Pizzas",
        "tipo": "Prepizza",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "CREPES VYP": {
        "categoria": "Pastas frescas",
        "tipo": "Crepes",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 4,
        "modificador_default": "Verdura y Pollo",
    },
    "CREPES PUERRO": {
        "categoria": "Pastas frescas",
        "tipo": "Crepes",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 4,
        "modificador_default": "Puerro, crema, pancetta y Champignon",
    },
    "PANQUEQUES": {
        "categoria": "Otros",
        "tipo": "Panqueques",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "TAP,EMPANADA": {
        "categoria": "Empanadas",
        "tipo": "Tapas de empanada",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "PASCUALINA": {
        "categoria": "Otros",
        "tipo": "Tapas pascualina",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "TAPA LASAGNA": {
        "categoria": "Otros",
        "tipo": "Tapa lasagna",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_KILO",
        "cantidad_default": 500,
    },
    "POSTRES": {
        "categoria": "Otros",
        "tipo": "Postres",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    # Pizzas
    "PIZA ESPECIAL": {
        "categoria": "Pizzas",
        "tipo": "Pizza Grande",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Especial",
    },
    "PIZA MUZA": {
        "categoria": "Pizzas",
        "tipo": "Pizza Grande",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Muzzarella",
    },
    "PIZA GOURMET": {
        "categoria": "Pizzas",
        "tipo": "Pizza Grande",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Gourmet",
    },
    "MIDI PIZZA MOZZARELLA": {
        "categoria": "Pizzas",
        "tipo": "Midi Pizza",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Muzzarella",
    },
    "MIDI PIZZA ESPECIAL": {
        "categoria": "Pizzas",
        "tipo": "Midi Pizza",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Especial",
    },
    "MIDI PIZZA GOURMET": {
        "categoria": "Pizzas",
        "tipo": "Midi Pizza",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
        "modificador_default": "Gourmet",
    },
    # Tartas
    "TARTA DE VERDURA GRANDE": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Grande",
        "modificador_default": "Verdura",
    },
    "TARTA DE VERDURA CHICA": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Chica",
        "modificador_default": "Verdura",
    },
    "TARTA TRICOLOR GRANDE": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Grande",
        "modificador_default": "Tricolor",
    },
    "TARTA TRICOLOR CHICA": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Chica",
        "modificador_default": "Tricolor",
    },
    "TARTA DE PUERRO GRANDE": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Grande",
        "modificador_default": "Puerro",
    },
    "TARTA DE PUERRO CHICA": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Chica",
        "modificador_default": "Puerro",
    },
    "TARTA  DE J&Q GRANDE": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Grande",
        "modificador_default": "Jamón y Queso",
    },
    "TARTA  DE J&Q CHICA": {
        "categoria": "Tartas",
        "tipo": "Tarta",
        "tamaño": "Chica",
        "modificador_default": "Jamón y Queso",
    },
    # Salsas
    "SS FILETTO": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Filetto"},
    "SS  BOLOG": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Bolognesa"},
    "SS RQUE Y 4 Q": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Roquefort y 4 Quesos"},
    "SS PPE": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Pomodoro"},
    "SS CREMA DE HONGOS": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Crema de Hongos"},
    "SS VERDEO": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Verdeo"},
    "SS PESTO": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Pesto"},
    "SS BCA": {"categoria": "Salsas", "tipo": "Salsa", "nombre": "Salsa Blanca"},
    "EMPANADAS": {
        "categoria": "Empanadas",
        "tipo": "Empanadas",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "PICADA": {
        "categoria": "Otros",
        "tipo": "Picada",
        "forma_venta": "UNIDAD",
        "unidad_precio": "POR_UNIDAD",
        "cantidad_default": 1,
    },
    "QUESO": {"categoria": "Otros", "tipo": "Queso", "nombre": "Queso (unidad)"},
}


# Productos en columna "Porciones calientes" (col 4-6 de Hoja 1)
MAPEO_PORCIONES_HOJA1: dict[str, dict[str, Any]] = {
    "RAV. SIMPLE": {"tipo": "Ravioles porción simple"},
    "FIDEOS SIMPLE": {"tipo": "Fideos porción simple"},
    "FORATTI / FUCCILE": {"tipo": "Pasta seca porción simple"},
    "FORATTI/FUCCILE/MOSTACHOLES": {"tipo": "Pasta seca porción simple"},
    "TORT. SIMPL": {"tipo": "Tortelettis porción simple"},
    "SORR SIMPLE": {"tipo": "Sorrentinos porción simple"},
    "LASAG SIMPLE": {"tipo": "Lasagna porción simple"},
    "RONDELLI SIMPLE": {"tipo": "Rondelli porción simple"},
    "CAN. SIMPLE": {"tipo": "Canelones porción simple"},
    "ÑOQUIS SIMPLE": {"tipo": "Ñoquis porción simple"},
    "CREPES SIMPLE": {"tipo": "Crepes porción simple"},
    "SORR NEGROS": {"tipo": "Sorrentinos negros porción"},
    "RAV, ESPECIAL": {"tipo": "Ravioles porción especial"},
    "FIDEOS ESPECIAL": {"tipo": "Fideos porción especial"},
    "FORATTI": {"tipo": "Foratti porción"},
    "TORTELETIS ESP": {"tipo": "Tortelettis porción especial"},
    "SORR ESPECIAL": {"tipo": "Sorrentinos porción especial"},
    "RAVIOLONES REMOLACHA": {"tipo": "Raviolones remolacha porción"},
    "LASAGNA ESPEC": {"tipo": "Lasagna porción especial"},
    "RONDELLI ESPEC": {"tipo": "Rondelli porción especial"},
    "CANELONES ESP": {"tipo": "Canelones porción especial"},
    "ÑOQUIS ESPEC": {"tipo": "Ñoquis porción especial"},
    "CREPES ESPEC": {"tipo": "Crepes porción especial"},
}


def normalizar(s: str | None) -> str:
    if not s:
        return ""
    return s.strip().upper().replace("Ñ", "Ñ").replace("\xc3‘", "Ñ").replace("�", "Ñ")


def fmt_precio(v: Any) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return ""


def parse_hoja_precios(ws) -> tuple[list[dict], list[dict]]:
    """Parsea Hoja 1: 3 columnas de productos con precios actuales."""
    productos: list[dict] = []
    porciones: list[dict] = []
    max_r = ws.max_row if ws.max_row else 100
    for row in ws.iter_rows(min_row=2, max_row=max_r, values_only=True):
        # Col 0-1: pastas
        nombre_a = normalizar(row[0]) if len(row) > 0 else ""
        precio_a = row[1] if len(row) > 1 else None
        if nombre_a and precio_a:
            mapping = MAPEO_PRODUCTOS_HOJA1.get(nombre_a)
            if mapping:
                productos.append({"raw": nombre_a, "precio": fmt_precio(precio_a), **mapping})
        # Col 2-3: pizzas/tartas
        nombre_b = normalizar(row[2]) if len(row) > 2 else ""
        precio_b = row[3] if len(row) > 3 else None
        if nombre_b and precio_b:
            mapping = MAPEO_PRODUCTOS_HOJA1.get(nombre_b)
            if mapping:
                productos.append({"raw": nombre_b, "precio": fmt_precio(precio_b), **mapping})
        # Col 4 (etiqueta) + 6 (precio): porciones calientes
        nombre_c = normalizar(row[4]) if len(row) > 4 else ""
        precio_c = row[6] if len(row) > 6 else None
        if nombre_c and precio_c:
            mapping = MAPEO_PORCIONES_HOJA1.get(nombre_c)
            if mapping:
                porciones.append(
                    {
                        "raw": nombre_c,
                        "precio": fmt_precio(precio_c),
                        "categoria": "Porciones calientes",
                        "forma_venta": "PORCION",
                        "unidad_precio": "POR_PORCION",
                        "cantidad_default": 1,
                        "cocina_interviene": True,
                        **mapping,
                    }
                )
    return productos, porciones


def parse_resto_simple(ws) -> dict[str, list[str]]:
    """Parsea hoja RESTO SIMPLE → mapping tipo → lista de opciones de relleno."""
    modificadores: dict[str, list[str]] = {}
    tipo_actual: str | None = None
    max_r = ws.max_row if ws.max_row else 100
    for row in ws.iter_rows(min_row=2, max_row=max_r, values_only=True):
        col0 = (row[0] or "").strip() if len(row) > 0 and row[0] else ""
        col1 = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        if col0:
            # Nuevo tipo
            tipo_actual = col0
            if tipo_actual not in modificadores:
                modificadores[tipo_actual] = []
        if col1 and tipo_actual:
            modificadores[tipo_actual].append(col1)
    return modificadores


def parse_pedidos_ya(ws) -> list[dict]:
    """Parsea hoja Pedidos YA — precios alternativos con recargo."""
    out: list[dict] = []
    max_r = ws.max_row if ws.max_row else 100
    for row in ws.iter_rows(min_row=2, max_row=max_r, values_only=True):
        nombre = normalizar(row[0]) if len(row) > 0 else ""
        precio = row[1] if len(row) > 1 else None
        if nombre and precio:
            try:
                p = float(precio)
                if p > 0:
                    out.append({"nombre": nombre, "precio_pedidos_ya": fmt_precio(p)})
            except (TypeError, ValueError):
                pass
    return out


def construir_seed(
    productos: list[dict],
    porciones: list[dict],
    modificadores_raw: dict[str, list[str]],
    pedidos_ya: list[dict],
) -> dict:
    """
    Construye la estructura final de seed con DEDUPLICACIÓN agresiva:
      - Agrupa productos por (categoria, tipo_nombre, [tamaño]) → 1 producto canónico.
      - Sabores con mismo precio → opciones del modificador con delta=0.
      - Sabores con precios distintos → opciones con delta_precio (relativo al menor).
      - Sorrentinos de Salmón se mantiene como producto separado (precio muy distinto).
    """
    todos = productos + porciones

    # ─── Agrupar por clave canónica ───────────────────────────────────────
    # Para tartas, distinguimos Grande/Chica como tipos separados.
    # Para Sorrentinos de Salmón, lo separamos del resto.
    def clave_canonica(p: dict) -> tuple[str, str]:
        cat = p.get("categoria", "Otros")
        tipo = p.get("tipo") or p.get("nombre", "Otros")
        # Tartas: separar por tamaño
        if cat == "Tartas" and p.get("tamaño"):
            tipo = f"Tarta {p['tamaño'].lower()}"
        # Sorrentinos de Salmón: producto separado
        if p.get("es_producto_separado"):
            tipo = p.get("nombre_separado", tipo)
        return (cat, tipo)

    grupos: dict[tuple[str, str], list[dict]] = {}
    for p in todos:
        k = clave_canonica(p)
        grupos.setdefault(k, []).append(p)

    # ─── Tipos de producto canónicos ──────────────────────────────────────
    tipos_canonicos = {}
    for (cat, tipo_nombre), items in grupos.items():
        # Si todos tienen cocina_interviene=true, hereda; si alguno lo tiene, el tipo lo tiene.
        cocina = any(it.get("cocina_interviene", False) for it in items)
        tipos_canonicos[(cat, tipo_nombre)] = {
            "categoria": cat,
            "nombre": tipo_nombre,
            "cocina_interviene": cocina,
            "descripcion": "",
        }

    # ─── Asignar códigos secuenciales por categoría (4 dígitos) ──────────
    # Categoría → rango:
    #   Pastas frescas: 0001-0999
    #   Pizzas:         1000-1499
    #   Tartas:         1500-1999
    #   Salsas:         2000-2499
    #   Empanadas:      2500-2999
    #   Porciones:      3000-3999
    #   Otros:          4000-4999
    rangos_categoria = {
        "Pastas frescas": 1,
        "Pizzas": 1000,
        "Tartas": 1500,
        "Salsas": 2000,
        "Empanadas": 2500,
        "Porciones calientes": 3000,
        "Otros": 4000,
    }
    contadores: dict[str, int] = {k: v for k, v in rangos_categoria.items()}

    # ─── Productos canónicos + modificadores con deltas ──────────────────
    productos_out = []
    modificadores_extra: list[dict] = []  # modificadores que se generan desde precios distintos

    for (cat, tipo_nombre), items in grupos.items():
        # Filtrar items con precio válido
        items_con_precio = [it for it in items if it.get("precio")]
        if not items_con_precio:
            continue

        # Tomar el primero como base, ordenado por precio
        items_sorted = sorted(items_con_precio, key=lambda x: float(x.get("precio") or 0))
        base = items_sorted[0]

        forma_venta = base.get("forma_venta", "UNIDAD")
        unidad_precio = base.get("unidad_precio", "POR_UNIDAD")
        cantidad_default = base.get("cantidad_default")
        precio_base = base.get("precio", "0.00")

        # Asignar código secuencial
        codigo_num = contadores.get(cat, 4000)
        contadores[cat] = codigo_num + 1
        codigo = f"{codigo_num:04d}"

        # Nombre canónico (sin sabor)
        nombre = tipo_nombre

        # Determinar sabores disponibles (modificadores)
        sabores: list[tuple[str, float]] = []  # (nombre, delta)
        if len(items_sorted) > 1:
            # Hay variantes con precios. Construimos lista de sabores.
            for it in items_sorted:
                sabor = it.get("modificador_default")
                if not sabor:
                    continue
                delta = float(it.get("precio", 0)) - float(precio_base)
                sabores.append((sabor, delta))
        elif base.get("modificador_default"):
            # 1 solo sabor en Hoja 1 — lo agregamos por las dudas, pero lo importante
            # es que el modificador real venga de RESTO SIMPLE.
            sabores.append((base["modificador_default"], 0.0))

        productos_out.append(
            {
                "codigo": codigo,
                "tipo_categoria": cat,
                "tipo_nombre": tipo_nombre,
                "nombre": nombre,
                "forma_venta": forma_venta,
                "unidad_precio": unidad_precio,
                "precio_base": precio_base,
                "cantidad_default": cantidad_default,
                "raw_excel_keys": [it.get("raw") for it in items],
            }
        )

        # Si hubo sabores con precios distintos detectados desde Hoja 1, generamos
        # un grupo de modificadores extra (que después se mergea con los de RESTO SIMPLE).
        if len(sabores) > 1:
            modificadores_extra.append(
                {
                    "tipo_producto": tipo_nombre,
                    "grupo_nombre": "Sabor",
                    "tipo_seleccion": "UNICA",
                    "obligatorio": True,
                    "opciones": [
                        {"nombre": s, "delta_precio": f"{d:.2f}"} for s, d in sabores
                    ],
                    "_origen": "hoja1_precios",
                }
            )

    # ─── Modificadores desde RESTO SIMPLE (sabores sin delta de precio) ──
    map_resto_simple = {
        "RAVIOLES": "Ravioles",
        "FIDEOS al HUEVO": "Fideos al huevo",
        "FIDEOS al MORRON": "Fideos al morrón",
        "FIDEOS VERDES (con ESPINACA)": "Fideos verdes",
        "ÑOQUIS de Semola": "Ñoquis de sémola",
        "SORRENTINOS": "Sorrentinos",
        "RAVIOLONES": "Raviolones",
        "LASAGNA": "Lasagna",
        "RONDELLI": "Rondelli",
        "CANELONES": "Canelones",
        "CREPES": "Crepes",
        "PIZZA GRANDE": "Pizza Grande",
    }

    modificadores_out: list[dict] = []
    tipos_con_mod_extra = {m["tipo_producto"] for m in modificadores_extra}

    # Primero agregar los desde RESTO SIMPLE (sin deltas)
    for raw, opciones in modificadores_raw.items():
        clave = map_resto_simple.get(raw)
        if not clave or not opciones:
            continue
        # Si ya hay un mod desde Hoja 1 (con deltas), tomamos esos sabores como base
        # y solo agregamos los que falten
        existente = next((m for m in modificadores_extra if m["tipo_producto"] == clave), None)
        if existente:
            # Mergear: tomar opciones de Hoja 1 (con delta) y agregar las de RESTO SIMPLE
            # que no estén
            nombres_existentes = {o["nombre"].lower() for o in existente["opciones"]}
            for o in opciones:
                if o.lower() not in nombres_existentes:
                    existente["opciones"].append({"nombre": o, "delta_precio": "0.00"})
            modificadores_out.append(existente)
            tipos_con_mod_extra.discard(clave)
        else:
            modificadores_out.append(
                {
                    "tipo_producto": clave,
                    "grupo_nombre": "Sabor",
                    "tipo_seleccion": "UNICA",
                    "obligatorio": True,
                    "opciones": [
                        {"nombre": o, "delta_precio": "0.00"} for o in opciones
                    ],
                }
            )

    # Agregar los modificadores de Hoja 1 que no se mergearon (no tienen entrada en RESTO SIMPLE)
    for m in modificadores_extra:
        if m["tipo_producto"] in tipos_con_mod_extra:
            modificadores_out.append(m)

    return {
        "categorias": CATEGORIAS,
        "tipos_producto": list(tipos_canonicos.values()),
        "productos": productos_out,
        "modificadores": modificadores_out,
        "precios_pedidos_ya": pedidos_ya,
        "_meta": {
            "total_productos": len(productos_out),
            "total_tipos": len(tipos_canonicos),
            "total_modificadores": len(modificadores_out),
            "fuente": "Lista de Precios.xlsx — hoja Hoja 1 (precios 16/04/2026) + RESTO SIMPLE",
            "version": 2,
            "deduplicado": True,
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Ruta a Lista de Precios.xlsx")
    ap.add_argument("--output", required=True, help="Archivo JSON de salida")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Leyendo {excel_path}...", file=sys.stderr)
    wb = load_workbook(excel_path, data_only=True)

    productos: list[dict] = []
    porciones: list[dict] = []
    modificadores_raw: dict[str, list[str]] = {}
    pedidos_ya: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue
        upper = sheet_name.upper()
        if upper == "HOJA 1":
            p, pc = parse_hoja_precios(ws)
            productos.extend(p)
            porciones.extend(pc)
        elif upper == "RESTO SIMPLE":
            modificadores_raw = parse_resto_simple(ws)
        elif "PEDIDOS YA" in upper:
            pedidos_ya = parse_pedidos_ya(ws)

    seed = construir_seed(productos, porciones, modificadores_raw, pedidos_ya)
    out_path.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")

    meta = seed["_meta"]
    print(f"\n✓ Seed generado: {out_path}", file=sys.stderr)
    print(f"  - Categorías:    {len(seed['categorias'])}", file=sys.stderr)
    print(f"  - Tipos:         {meta['total_tipos']}", file=sys.stderr)
    print(f"  - Productos:     {meta['total_productos']}", file=sys.stderr)
    print(f"  - Modificadores: {meta['total_modificadores']}", file=sys.stderr)
    print(f"  - Pedidos YA:    {len(pedidos_ya)}", file=sys.stderr)


if __name__ == "__main__":
    main()
