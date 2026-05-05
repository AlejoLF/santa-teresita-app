#!/usr/bin/env python3
"""
Parser de "Proveedores 2026.xlsx" — hojas Estantería y Bebidas.
Genera JSON de seed con productos de estantería (envasados) y bebidas (vinos, aguas, cervezas).

Uso:
    python tools/parse_estanteria_bebidas.py \
        --excel "Proveedores 2026.xlsx" \
        --output packages/db/prisma/seed-data/estanteria-bebidas.json

Convenciones de las hojas:
  Estantería:
    R1 cabecera: Concepto | Presentacion | Marca | Precio | x 1.6 | Se vende a
    Filas siguientes: 1 producto por fila. La columna "Se vende a" es el precio final.
    Sub-sección "Pastas Troncoso" desde R203 — primer columna es la VARIEDAD
    (SEMOLA / MULTIGRANO / 3 SABORES / ESPINACA / CALAMAR / ZANAHORIA), col2 es la FORMA
    (Conchiglioni / Fusilloni / Penne Rigate / Maccarun Rigatti). La variedad
    se hereda hacia abajo cuando la celda está vacía.

  Bebidas:
    R1 cabecera: VINOS | Uva | Costo | Precio | de Venta
    Filas: producto | variedad | costo | precio_calc | precio_final
    No tiene marca explícita — el "Concepto" es el nombre del vino, que en muchos
    casos es la marca también.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Falta openpyxl. Instalá con: pip install openpyxl\n")
    sys.exit(1)


# Códigos por categoría — coordinados con parse_lista_precios.py
CODIGO_BASE_ESTANTERIA = 5000
CODIGO_BASE_BEBIDAS = 6000


def normalizar(s):
    if s is None:
        return ""
    return str(s).strip()


def fmt_precio(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
        if n <= 0:
            return None
        return f"{n:.2f}"
    except (TypeError, ValueError):
        return None


# ────────────────────────────────────────────────────────────────────────
#   ESTANTERÍA
# ────────────────────────────────────────────────────────────────────────


def parse_estanteria(ws):
    """Devuelve productos planos de estantería + sección de pastas Troncoso."""
    productos = []
    pastas = []  # productos de pastas Troncoso, manejados aparte
    en_seccion_pastas = False
    variedad_actual = None  # SEMOLA / MULTIGRANO / etc.

    max_r = ws.max_row if ws.max_row else 1200
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_r, values_only=True), start=2):
        # row[0] concepto | row[1] presentacion | row[2] marca | row[3] precio costo |
        # row[4] x 1.6 | row[5] se vende a (precio final)
        concepto = normalizar(row[0]) if len(row) > 0 else ""
        presentacion = normalizar(row[1]) if len(row) > 1 else ""
        marca = normalizar(row[2]) if len(row) > 2 else ""
        precio_final = fmt_precio(row[5]) if len(row) > 5 else None

        # Detectar inicio de sección Pastas Troncoso por la marca "Troncoso"
        if marca == "Troncoso":
            en_seccion_pastas = True
            # En esta sección, la columna concepto puede ser la VARIEDAD
            # (SEMOLA, MULTIGRANO, 3 SABORES, ESPINACA, CALAMAR, ZANAHORIA)
            # o estar vacía (heredando la última variedad)
            if concepto:
                variedad_actual = concepto
            forma = presentacion  # en pastas, "presentación" es la forma física
            if precio_final and forma and variedad_actual:
                pastas.append(
                    {
                        "variedad": variedad_actual,
                        "forma": forma,
                        "marca": "Troncoso",
                        "precio": precio_final,
                    }
                )
            continue

        # Si estábamos en pastas y ya no estamos, salimos.
        if en_seccion_pastas and marca and marca != "Troncoso":
            en_seccion_pastas = False

        # Producto regular
        if concepto and precio_final:
            productos.append(
                {
                    "concepto": concepto,
                    "presentacion": presentacion or None,
                    "marca": marca or None,
                    "precio": precio_final,
                }
            )

    return productos, pastas


# ────────────────────────────────────────────────────────────────────────
#   BEBIDAS
# ────────────────────────────────────────────────────────────────────────


def parse_bebidas(ws):
    """Devuelve lista plana de bebidas con concepto (=marca/nombre) + variedad + precio."""
    bebidas = []
    max_r = ws.max_row if ws.max_row else 50
    for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=max_r, values_only=True), start=3):
        # row[0] Concepto/Marca | row[1] Uva/Variedad | row[2] Costo | row[3] Precio | row[4] de Venta
        concepto = normalizar(row[0]) if len(row) > 0 else ""
        variedad = normalizar(row[1]) if len(row) > 1 else ""
        precio_final = fmt_precio(row[4]) if len(row) > 4 else None
        if not concepto or not precio_final:
            continue
        bebidas.append(
            {
                "concepto": concepto,
                "variedad": variedad or None,
                "precio": precio_final,
            }
        )
    return bebidas


# ────────────────────────────────────────────────────────────────────────
#   Construcción del seed JSON
# ────────────────────────────────────────────────────────────────────────


def construir_seed(estanteria, pastas_troncoso, bebidas):
    """Genera un JSON con productos canónicos para estantería y bebidas."""
    # Categorías nuevas
    categorias_extra = [
        {"nombre": "Estantería", "orden": 8, "icono": "🧴"},
        {"nombre": "Bebidas", "orden": 9, "icono": "🍷"},
    ]

    # ─── Tipos de producto ─────────────────────────────────────────────
    tipos_producto = []
    tipos_set = set()

    def agregar_tipo(cat, nombre, cocina=False):
        key = (cat, nombre)
        if key in tipos_set:
            return
        tipos_set.add(key)
        tipos_producto.append(
            {
                "categoria": cat,
                "nombre": nombre,
                "cocina_interviene": cocina,
                "descripcion": "",
            }
        )

    # ─── Productos ──────────────────────────────────────────────────────
    productos = []
    contador_estanteria = CODIGO_BASE_ESTANTERIA
    contador_bebidas = CODIGO_BASE_BEBIDAS

    # 1. Estantería regular: agrupamos por marca como "tipo" para que la subsección
    #    en la UI sea clara. Tipo = marca (cuando hay), sino "Sin marca".
    #    Ej: tipo "De La Torre" agrupa todos los productos de De La Torre.
    for p in estanteria:
        marca = p["marca"] or "Sin marca"
        # Tipo del producto = marca
        agregar_tipo("Estantería", marca, cocina=False)

        codigo = f"{contador_estanteria:04d}"
        contador_estanteria += 1
        productos.append(
            {
                "codigo": codigo,
                "tipo_categoria": "Estantería",
                "tipo_nombre": marca,
                "nombre": p["concepto"],
                "marca": p["marca"],
                "presentacion": p["presentacion"],
                "forma_venta": "UNIDAD",
                "unidad_precio": "POR_UNIDAD",
                "precio_base": p["precio"],
                "cantidad_default": 1,
            }
        )

    # 2. Pastas Troncoso — sub-sección dedicada
    agregar_tipo("Estantería", "Pastas secas (Troncoso)", cocina=False)
    for p in pastas_troncoso:
        codigo = f"{contador_estanteria:04d}"
        contador_estanteria += 1
        nombre = f"{p['variedad'].title()} {p['forma']}"  # "Semola Conchiglioni"
        productos.append(
            {
                "codigo": codigo,
                "tipo_categoria": "Estantería",
                "tipo_nombre": "Pastas secas (Troncoso)",
                "nombre": nombre,
                "marca": p["marca"],
                "presentacion": None,
                "forma_venta": "UNIDAD",
                "unidad_precio": "POR_UNIDAD",
                "precio_base": p["precio"],
                "cantidad_default": 1,
            }
        )

    # 3. Bebidas: agrupamos por concepto+variedad-tipo. Para vinos típicos,
    #    el tipo = primera palabra para agrupar el catálogo (ej. todos los Malbec).
    #    Más simple: tipo = "Vinos" / "Aguas" / "Cervezas" / "Otras bebidas"
    def tipo_bebida(p):
        c = p["concepto"].lower()
        v = (p["variedad"] or "").lower()
        if "agua" in c:
            return "Aguas"
        if "andes" in c or "cerveza" in c or "ipa" in v:
            return "Cervezas"
        # Resto se considera vino
        return "Vinos"

    for p in bebidas:
        tipo = tipo_bebida(p)
        agregar_tipo("Bebidas", tipo, cocina=False)
        codigo = f"{contador_bebidas:04d}"
        contador_bebidas += 1
        nombre = p["concepto"]
        if p["variedad"]:
            nombre = f"{p['concepto']} ({p['variedad']})"
        productos.append(
            {
                "codigo": codigo,
                "tipo_categoria": "Bebidas",
                "tipo_nombre": tipo,
                "nombre": nombre,
                "marca": p["concepto"],  # el concepto es típicamente la marca
                "presentacion": None,
                "forma_venta": "UNIDAD",
                "unidad_precio": "POR_UNIDAD",
                "precio_base": p["precio"],
                "cantidad_default": 1,
            }
        )

    return {
        "categorias": categorias_extra,
        "tipos_producto": tipos_producto,
        "productos": productos,
        "_meta": {
            "total_productos": len(productos),
            "total_estanteria": sum(1 for p in productos if p["tipo_categoria"] == "Estantería"),
            "total_bebidas": sum(1 for p in productos if p["tipo_categoria"] == "Bebidas"),
            "total_pastas_troncoso": len(pastas_troncoso),
            "fuente": "Proveedores 2026.xlsx — hojas Estantería + Bebidas (precio columna 'Se vende')",
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    excel_path = Path(args.excel)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Leyendo {excel_path}...", file=sys.stderr)
    wb = load_workbook(excel_path, data_only=True)

    if "Estanteria" not in wb.sheetnames:
        sys.stderr.write("✕ Falta hoja 'Estanteria' en el Excel\n")
        sys.exit(1)
    if "Bebidas" not in wb.sheetnames:
        sys.stderr.write("✕ Falta hoja 'Bebidas' en el Excel\n")
        sys.exit(1)

    estanteria, pastas = parse_estanteria(wb["Estanteria"])
    bebidas = parse_bebidas(wb["Bebidas"])
    seed = construir_seed(estanteria, pastas, bebidas)

    out_path.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    meta = seed["_meta"]
    print(f"\n✓ Seed generado: {out_path}", file=sys.stderr)
    print(f"  - Productos:           {meta['total_productos']}", file=sys.stderr)
    print(f"  - Estantería:          {meta['total_estanteria']}", file=sys.stderr)
    print(f"  - Bebidas:             {meta['total_bebidas']}", file=sys.stderr)
    print(f"  - Pastas Troncoso:     {meta['total_pastas_troncoso']}", file=sys.stderr)
    print(f"  - Tipos de producto:   {len(seed['tipos_producto'])}", file=sys.stderr)


if __name__ == "__main__":
    main()
